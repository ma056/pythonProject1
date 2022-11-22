# -- coding: utf-8 --
"""
@Project : pythonProject1
@File : openpyxl1.py
@Author : wenjing
@Date : 2022/11/22 10:42

"""

'''
1.创建文件
from openpyxl import Workbook

# 创建一个excel文件
wb = Workbook()
# 创建一个子表sheet,0是设置子表的位置，目前是第一个
ws1 = wb.create_sheet('用户信息表', 0)
ws2 = wb.create_sheet('bank_info')
wb.remove('bank_info')  # 删除表的方法一
del wb('bank_info')  # 删除表的方法二
# 更改ws1的名称
ws1.title = 'user_info'
# 修改子表背景颜色
ws1.sheet_properties.tabColor = 'FF6666'

# 查看excel所有子表的信息
print(wb.sheetnames)
# 打印结果：['user_info','sheet','bank_info']
# Workbook.copy_worksheet(表名) —创建工作表的副本

# 访问一个单元格并赋值
ws1['A1'] = 'wenjing'

# Worksheet.cell() 此方法是通过指定坐标的方式赋值，其中 row 参数代表行数， column 代表列数，value 是填入的数据
ws1.cell(row=4, column=2, value=10)

# Workbook.append() 同时写入多个数据
ws1.append(['name', 'age', 'gender'])
ws1.append(['wenjing', '15', 'man'])
for row in ws1.values:
    print(row)
# 打印结果：('name','age', 'gender')
#         ('wenjing', '15', 'man')
# 保存文件
wb = Workbook()
wb.save('balances.xlsx')  # balances.xlsx 是保存的路径，也就是文件名。

# 补充：
ws1['A6'] = '=sum(A4:A5)'  # 对A4，A5进行求和并写入A6
'''
from openpyxl import Workbook

# 创建一个excel文件
wb = Workbook()
# 创建一个子表sheet,0是设置子表的位置，目前是第一个
ws1 = wb.create_sheet('用户信息表', 0)
ws2 = wb.create_sheet('bank_info')
wb.remove('bank_info')  # 删除表的方法一
# del wb('bank_info')  # 删除表的方法二
# 更改ws1的名称
ws1.title = 'user_info'
# 修改子表背景颜色
ws1.sheet_properties.tabColor = 'FF6666'

# 查看excel所有子表的信息
print(wb.sheetnames)
# 打印结果：['user_info','sheet','bank_info']
# Workbook.copy_worksheet() —创建工作表的副本

# 访问一个单元格并赋值
ws1['A1'] = 'wenjing'

# Worksheet.cell() 此方法是通过指定坐标的方式赋值，其中 row 参数代表行数， column 代表列数，value 是填入的数据
ws1.cell(row=4, column=2, value=10)

# Workbook.append() 同时写入多个数据
ws1.append(['name', 'age', 'gender'])
ws1.append(['wenjing', '15', 'man'])
for row in ws1.values:
    print(row)
# 打印结果：('name','age', 'gender')
#         ('wenjing', '15', 'man')
# 保存文件
wb = Workbook()
wb.save('balances.xlsx')  # balances.xlsx 是保存的路径，也就是文件名。

# 补充：
ws1['A6'] = '=sum(A4:A5)'  # 对A4，A5进行求和并写入A6

'''
2.读取文件
'''
from openpyxl import load_workbook

wb = load_workbook('data.xlsx', read_only=True, data_only=True)
print(wb.sheetnames)  # 打印所有子表

# 读取子表数据
# 方式一：
wb = load_workbook('data.xlsx', read_only=True, data_only=True)
ws = wb['用户信息表']
print(ws['A1'].value)  # 获取子表中A1的位置

# 方式二：
wb = load_workbook('data.xlsx', read_only=True, data_only=True)
ws = wb['用户信息表']
print(ws.cell(row=2, column=1).value)  # 通过行与列的形式来取值

# 方式三：
wb = load_workbook('ex_a.xlsx', read_only=True, data_only=True)
ws = wb['用户信息表']
for row in ws.rows:  # 获取每一行的值
    for data in row:  # 获取一行中单元格的数据
        print(data.value)  # 打印单元格的值

# 单元格复制
wb = load_workbook('ex_a.xlsx', read_only=True, data_only=True)
ws = wb['用户信息表']
ws['A1'] = '第一行第一列修改过后的单元格值'
ws.cell(1, 2).value = '第一行第二列修改过后的单元格值'

# 指定sheet表种删除行和删除列
wb = load_workbook('ex_a.xlsx', read_only=True, data_only=True)
ws = wb['用户信息表']
ws.delete_cols(1)  # 删除第一列，以此类推、n代表删除第n列
ws.delete_rows(1)  # 删除第一行，以此类推、n代表删除第n行

# 合并的单元格
wb = load_workbook('ex_a.xlsx', read_only=True, data_only=True)
ws = wb['用户信息表']
ws.merge_cells("A1:B1")
ws.merge_cells(start_column = 3,end_column=6,start_row=2,end_row = 3)